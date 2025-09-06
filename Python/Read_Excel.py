import openpyxl

# Fetch Whole Excel
def Fetch_Whole_Excel(SheetPath):
    book = openpyxl.load_workbook(SheetPath)
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)
        print("--------")

    book.close()

# Fetch Specific row value
def Fetch_data_basedon_TC(SheetPath, TC):
    book = openpyxl.load_workbook(SheetPath)
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)

    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == TC:
            for j in range(1, sheet.max_column + 1):
                print(sheet.cell(row=i, column=j).value)

    book.close()

def Fetch_and_store_in_Dict(SheetPath):
    book = openpyxl.load_workbook(SheetPath)
    sheet = book.active
    print(sheet.max_row)
    print(sheet.max_column)
    Dict = {}
    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value]= (sheet.cell(row=i, column=j).value)
        print(Dict)
    book.close()

def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}
 
    for i in range(1,sheet.max_column+ 1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"] = i
 
    for i in range(1,sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i,column= j).value == searchTerm:
                Dict["row"] = i
 
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(filePath)



# Calling Function's

# Fetch_data_basedon_TC(r"E:\Selenium\Selenium_Python\Documents\Openpyxl.xlsx", "TC03")
# Fetch_and_store_in_Dict(r"E:\Selenium\Selenium_Python\Documents\Openpyxl.xlsx")
# Fetch_Whole_Excel(r"E:\Selenium\Selenium_Python\Documents\Openpyxl.xlsx")
 
